import json
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

def parse_machine(file_path):
    """Parse NDJSON Flutter test machine output"""
    tests = []
    try:
        with open(file_path, 'r', encoding='utf-8-sig') as f:
            content = f.read().strip()
            if not content:
                return tests
            
            # Try parsing as single JSON object first
            try:
                obj = json.loads(content)
                if isinstance(obj, dict) and obj.get('type'):
                    tests.append(obj)
                    return tests
            except json.JSONDecodeError:
                pass
            
            # Parse as newline-delimited JSON
            for line in content.split('\n'):
                line = line.strip()
                if not line or line.startswith('['):
                    continue
                try:
                    obj = json.loads(line)
                    if isinstance(obj, dict) and obj.get('type'):
                        tests.append(obj)
                except json.JSONDecodeError:
                    continue
    except Exception as e:
        print(f'Warning: Error reading {file_path}: {e}', file=sys.stderr)
    
    return tests

def extract_test_name(test_obj):
    """Extract test name from nested test object"""
    if isinstance(test_obj, dict):
        return test_obj.get('name', 'Unknown')
    return str(test_obj) if test_obj else 'Unknown'

def write_xlsx(events, out_path):
    """Write test events to XLSX with formatting"""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Test Results'
    
    # Header formatting
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    
    headers = ['Event Type', 'Time (ms)', 'Test Name', 'Result', 'Skipped', 'Status', 'Message']
    ws.append(headers)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Set column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 40
    
    if not events:
        ws.append(['No test events', '', '', '', '', '', 'No tests were run or results were empty'])
    else:
        # Track test counts
        passed_count = 0
        failed_count = 0
        
        for ev in events:
            event_type = ev.get('type', '')
            time = ev.get('time', '')
            
            # Extract test name from nested test object
            test_obj = ev.get('test', {})
            test_name = extract_test_name(test_obj)
            
            result = ev.get('result', '')
            skipped = ev.get('skipped', False)
            hidden = ev.get('hidden', False)
            message = ''
            
            # Determine status and message
            status = ''
            if event_type == 'testDone':
                status = result if result else 'completed'
                if result == 'success':
                    passed_count += 1
                elif result == 'error':
                    failed_count += 1
            elif event_type == 'testError':
                status = 'error'
                message = ev.get('error', '')
                failed_count += 1
            elif event_type == 'testStart':
                status = 'started'
            elif event_type == 'start':
                status = 'suite-started'
            elif event_type == 'done':
                status = 'suite-done'
            
            if ev.get('type') == 'print':
                message = ev.get('message', '')[:500]  # Limit message length
            
            ws.append([event_type, time, test_name, result, skipped, status, message])
        
        # Add summary row
        ws.append([])
        summary_row = ws.max_row + 1
        ws[f'A{summary_row}'] = 'Summary'
        ws[f'B{summary_row}'] = f'Passed: {passed_count}, Failed: {failed_count}'
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    wb.save(out_path)
    return len(events)

if __name__ == '__main__':
    if len(sys.argv) < 3:
        print('Usage: convert_flutter_test_machine_to_xlsx.py input.json output.xlsx')
        sys.exit(2)
    
    input_file = sys.argv[1]
    output_file = sys.argv[2]
    
    try:
        events = parse_machine(input_file)
        event_count = write_xlsx(events, output_file)
        print(f'Successfully wrote {output_file} with {event_count} test event(s)')
        sys.exit(0)
    except Exception as e:
        print(f'Error: {e}', file=sys.stderr)
        # Create empty report on error
        wb = Workbook()
        ws = wb.active
        ws.append(['Error', 'Failed to parse test results', str(e)])
        wb.save(output_file)
        print(f'Created error report in {output_file}')
        sys.exit(1)

