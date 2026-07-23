from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib import colors

metrics = {
    'Test Configuration': '',
    'Target': 'https://api.open-meteo.com/v1/forecast',
    'Virtual users': 100,
    'Duration (seconds)': 60,
    'Ramp-up (seconds)': 10,
    'Target RPS': 120,
    '': '',
    'Results': '',
    'Total requests': 2711,
    'Successful requests': 2673,
    'Failed requests': 38,
    'Error rate (%)': 1.40,
    '': '',
    'Throughput': '',
    'Requests per second (RPS)': 45.18,
    '': '',
    'Response Times': '',
    'Average response time (ms)': 1702.2,
    'Min response time (ms)': 1208.8,
    'Max response time (ms)': 6338.5,
    'p50 response time (ms)': 1601.8,
    'p90 response time (ms)': 2137.8,
    'p95 response time (ms)': 2434.5,
    'p99 response time (ms)': 3070.0,
    '': '',
    'Status Code Distribution': '',
    'HTTP 200 (Success)': 712,
    'HTTP 429 (Rate Limited)': 1961,
}

status_summary = (
    'Baseline load test with 100 VUs over 60 seconds (10s ramp-up, 120 RPS target). '
    'Achieved 45.18 RPS. Average response time: 1702.2 ms. Error rate: 1.40%. '
    'Finding: External API endpoint (api.open-meteo.com) is rate-limited. Out of 2673 successful HTTP responses, '
    '1961 (73%) were HTTP 429 (Too Many Requests), indicating the endpoint cannot handle the target throughput without rate limiting. '
    'Only 712 requests (27%) received HTTP 200 (Success). Average response time exceeds 300ms target (1702.2ms measured).'
)


def write_excel(path):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Load Test Status'

    title = 'Baseline Load Test Status'
    ws.merge_cells('A1:B1')
    ws['A1'] = title
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')

    # Header row
    header_fill = PatternFill(start_color='FFD966', end_color='FFD966', fill_type='solid')
    ws['A3'] = 'Metric'
    ws['B3'] = 'Value'
    for cell in (ws['A3'], ws['B3']):
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    row = 4
    for name, value in metrics.items():
        ws[f'A{row}'] = name
        ws[f'B{row}'] = value
        row += 1

    # Add prompt-style response time summary below
    row += 1
    ws[f'A{row}'] = 'Summary'
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'] = status_summary
    ws.row_dimensions[row].height = 30
    ws[f'B{row}'].alignment = Alignment(wrap_text=True)

    # Column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 90

    wb.save(path)


def write_pdf(path):
    styles = getSampleStyleSheet()
    story = []
    story.append(Paragraph('Baseline Load Test Status', styles['Heading1']))
    story.append(Spacer(1, 12))
    story.append(Paragraph(status_summary, styles['BodyText']))
    story.append(Spacer(1, 18))

    data = [['Metric', 'Value']]
    for name, value in metrics.items():
        data.append([name, str(value)])

    table = Table(data, colWidths=[220, 320])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
    ]))
    story.append(table)
    doc = SimpleDocTemplate(path, pagesize=letter)
    doc.build(story)


if __name__ == '__main__':
    write_excel('tooling/load_test_report.xlsx')
    write_pdf('tooling/load_test_report.pdf')
    print('Generated tooling/load_test_report.xlsx and tooling/load_test_report.pdf')
